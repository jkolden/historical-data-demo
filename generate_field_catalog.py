"""
Parse all DDL files and generate an Excel field catalog
with one sheet per table for team review.
"""
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))

# (sheet_label, ddl_file_path)
TABLES = [
    ("AP Invoices",       "ap-invoices/demo_ap_invoices_hist.sql"),
    ("AP Invoice Lines",  "ap-invoices/demo_ap_invoice_line_hist.sql"),
    ("Purchase Orders",   "purchase-orders/demo_po_hist_ddl.sql"),
    ("Assignments",       "assignments/demo_assignment_hist_ddl.sql"),
    ("Absences",          "absences/demo_absence_hist_ddl.sql"),
    ("AR Invoices",       "receivables/demo_ar_invoices_hist.sql"),
    ("Checks",            "checks/demo_checks_hist.sql"),
    ("Learning",          "learning/demo_learning_hist.sql"),
]

# Regex: captures column_name, datatype (with optional size), and optional NOT NULL / inline comment
COL_RE = re.compile(
    r'^\s+'
    r'(?!constraint\b|--)'                       # skip constraints and comments
    r'(\w+)'                                      # column name
    r'\s+'
    r'((?:number|varchar2|date|timestamp|clob)'   # datatype keyword
    r'(?:\s*\([^)]+\))?'                          # optional (size)
    r'(?:\s+with\s+time\s+zone)?)'                # optional WITH TIME ZONE
    r'(.*)',                                       # rest of line (constraints, comments)
    re.IGNORECASE
)

IDENTITY_RE = re.compile(r'generated\s+always\s+as\s+identity', re.IGNORECASE)
COMMENT_RE  = re.compile(r'--\s*(.*)')
SECTION_RE  = re.compile(r'^\s*--\s*=+\s*(.*?)\s*=*\s*$')

# Styles
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_FONT   = Font(name='Calibri', size=11)
CELL_ALIGN  = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
ALT_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')


def parse_ddl(filepath):
    """Return list of dicts: {name, datatype, nullable, pk, identity, section, comment}"""
    columns = []
    current_section = ""
    with open(filepath, encoding='utf-8') as f:
        for line in f:
            # Check for section header comment like "-- === Section Name ==="
            sec = SECTION_RE.match(line)
            if sec:
                current_section = sec.group(1).strip().rstrip('=').strip()
                # Also handle "---- Section ----" style
                current_section = current_section.strip('-').strip()
                continue

            # Also handle "-- ---- Section ----" style (AR invoices)
            sec2 = re.match(r'^\s*--\s*-+\s*(.*?)\s*-*\s*$', line)
            if sec2:
                txt = sec2.group(1).strip()
                if txt and len(txt) > 2:
                    current_section = txt
                continue

            m = COL_RE.match(line)
            if not m:
                continue

            col_name = m.group(1).lower()
            datatype = m.group(2).strip()
            rest     = m.group(3).strip().rstrip(',')

            # Normalize datatype display
            datatype = re.sub(r'\s+', ' ', datatype).upper()

            nullable = 'N' if 'NOT NULL' in rest.upper() else 'Y'
            identity = 'Y' if IDENTITY_RE.search(rest) else ''
            pk = 'Y' if 'PRIMARY KEY' in rest.upper() else ''

            # Extract inline comment
            cm = COMMENT_RE.search(rest)
            comment = cm.group(1).strip() if cm else ''

            columns.append({
                'name': col_name,
                'datatype': datatype,
                'nullable': nullable,
                'pk': pk,
                'identity': identity,
                'section': current_section,
                'comment': comment,
            })
    return columns


def build_workbook():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    headers = ['#', 'Column Name', 'Data Type', 'Nullable', 'PK', 'Identity', 'Section', 'Comment']
    col_widths = [5, 45, 30, 9, 5, 9, 30, 50]

    for sheet_label, ddl_rel in TABLES:
        ddl_path = os.path.join(BASE, ddl_rel)
        columns = parse_ddl(ddl_path)

        ws = wb.create_sheet(title=sheet_label)

        # Extract table name from DDL for subtitle
        with open(ddl_path, encoding='utf-8') as f:
            content = f.read()
        tbl_match = re.search(r'CREATE\s+TABLE\s+(\w+)', content, re.IGNORECASE)
        table_name = tbl_match.group(1).upper() if tbl_match else sheet_label

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=f"{table_name}  ({len(columns)} columns)")
        title_cell.font = Font(name='Calibri', bold=True, size=14, color='2F5496')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 28

        # Header row
        for ci, hdr in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=hdr)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        ws.row_dimensions[2].height = 22

        # Data rows
        for ri, col in enumerate(columns, 3):
            row_data = [
                ri - 2,
                col['name'],
                col['datatype'],
                col['nullable'],
                col['pk'],
                col['identity'],
                col['section'],
                col['comment'],
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGN
                cell.border = THIN_BORDER
                if (ri - 3) % 2 == 1:
                    cell.fill = ALT_FILL

        # Column widths
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=2, column=ci).column_letter].width = w

        # Freeze panes below header
        ws.freeze_panes = 'A3'

        # Auto-filter
        ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(headers)).column_letter}{2 + len(columns)}"

    # Summary sheet at the front
    summary = wb.create_sheet(title="Summary", index=0)
    summary.merge_cells('A1:D1')
    t = summary.cell(row=1, column=1, value="Historical Data Demo - Table Field Catalog")
    t.font = Font(name='Calibri', bold=True, size=16, color='2F5496')
    t.alignment = Alignment(horizontal='left')
    summary.row_dimensions[1].height = 30

    sum_headers = ['Table Name', 'Sheet', 'Column Count', 'REST Endpoint']
    for ci, h in enumerate(sum_headers, 1):
        c = summary.cell(row=3, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER

    endpoints = {
        "AP Invoices":      "/fscmRestApi/resources/11.13.18.05/invoices",
        "AP Invoice Lines": "/fscmRestApi/.../invoices/{id}/child/invoiceLines",
        "Purchase Orders":  "/fscmRestApi/resources/11.13.18.05/purchaseOrders",
        "Assignments":      "/hcmRestApi/resources/11.13.18.05/publicWorkers + child/assignments",
        "Absences":         "/hcmRestApi/resources/11.13.18.05/absences",
        "AR Invoices":      "/fscmRestApi/resources/11.13.18.05/receivablesInvoices",
        "Checks":           "/fscmRestApi/resources/11.13.18.05/payablesPayments",
        "Learning":         "/hcmRestApi/resources/11.13.18.05/learnerLearningRecords",
    }

    total_cols = 0
    for ri, (label, ddl_rel) in enumerate(TABLES, 4):
        ddl_path = os.path.join(BASE, ddl_rel)
        cols = parse_ddl(ddl_path)
        total_cols += len(cols)
        with open(ddl_path, encoding='utf-8') as f:
            content = f.read()
        tbl_match = re.search(r'CREATE\s+TABLE\s+(\w+)', content, re.IGNORECASE)
        tbl_name = tbl_match.group(1).upper() if tbl_match else label

        row_data = [tbl_name, label, len(cols), endpoints.get(label, '')]
        for ci, val in enumerate(row_data, 1):
            c = summary.cell(row=ri, column=ci, value=val)
            c.font = CELL_FONT
            c.alignment = CELL_ALIGN
            c.border = THIN_BORDER
            if (ri - 4) % 2 == 1:
                c.fill = ALT_FILL

    # Total row
    tr = 4 + len(TABLES)
    summary.cell(row=tr, column=1, value="TOTAL").font = Font(name='Calibri', bold=True, size=11)
    summary.cell(row=tr, column=3, value=total_cols).font = Font(name='Calibri', bold=True, size=11)
    for ci in range(1, 5):
        summary.cell(row=tr, column=ci).border = THIN_BORDER

    summary.column_dimensions['A'].width = 35
    summary.column_dimensions['B'].width = 20
    summary.column_dimensions['C'].width = 14
    summary.column_dimensions['D'].width = 60
    summary.freeze_panes = 'A4'

    return wb


if __name__ == '__main__':
    wb = build_workbook()
    out = os.path.join(BASE, 'Historical_Data_Field_Catalog.xlsx')
    wb.save(out)
    print(f"Saved: {out}")
    # Count sheets and columns
    for ws in wb.worksheets:
        print(f"  Sheet '{ws.title}': {ws.max_row - 2} rows" if ws.title != 'Summary' else f"  Sheet '{ws.title}'")
