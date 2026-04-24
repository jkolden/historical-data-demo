"""
Parse all REST API describe JSONs and generate an Excel catalog
of queryable fields by subject area / endpoint.
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))

# (sheet_label, describe_file, resource_key, child_keys_to_include)
ENDPOINTS = [
    ("AP Invoices",       "ap-invoices/invoicesDescribe.json",               "invoices",              []),
    ("AP Invoice Lines",  "ap-invoices/invoicesDescribe.json",               "invoices",              ["invoiceLines"]),
    ("AP Installments",   "ap-invoices/invoicesDescribe.json",               "invoices",              ["invoiceInstallments"]),
    ("Purchase Orders",   "purchase-orders/purchaseOrdersDescribe.json",     "purchaseOrders",        []),
    ("PO Lines",          "purchase-orders/purchaseOrdersDescribe.json",     "purchaseOrders",        ["lines"]),
    ("Assignments",       "assignments/publicWorkersDescribe.json",          "publicWorkers",         ["assignments"]),
    ("Workers",           "assignments/publicWorkersDescribe.json",          "publicWorkers",         []),
    ("Absences",          "absences/absencesDescribe.json",                  "absences",              []),
    ("AR Invoices",       "receivables/receivablesInvoicesDescribe.json",    "receivablesInvoices",   []),
    ("AR Invoice Lines",  "receivables/receivablesInvoicesDescribe.json",    "receivablesInvoices",   ["receivablesInvoiceLines"]),
    ("Checks",            "checks/payablesPaymentsDescribe.json",           "payablesPayments",      []),
    ("Learning Records",  "learning/learnerLearningRecordsDescribe.json",   "learnerLearningRecords",[]),
]

# Styles
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_FONT   = Font(name='Calibri', size=11)
CELL_ALIGN  = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
ALT_FILL    = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
YES_FONT    = Font(name='Calibri', size=11, color='2E7D32')
NO_FONT     = Font(name='Calibri', size=11, color='C62828')


def extract_attrs(describe_path, resource_key, child_keys):
    """Extract attributes from a describe JSON. If child_keys is empty,
    return parent-level attrs. If child_keys is specified, return those
    child resource attrs."""
    with open(describe_path, encoding='utf-8') as f:
        data = json.load(f)

    resource = data.get('Resources', {}).get(resource_key, {})

    if not child_keys:
        # Parent-level attributes
        return resource_key, resource.get('attributes', [])
    else:
        # Child-level attributes
        children = resource.get('children', {})
        all_attrs = []
        label = child_keys[0]
        for ck in child_keys:
            child = children.get(ck, {})
            all_attrs.extend(child.get('attributes', []))
        return label, all_attrs


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    headers = ['#', 'Field Name', 'Type', 'Queryable', 'Updatable', 'Mandatory',
               'Max Length', 'Title', 'Description']
    col_widths = [5, 45, 12, 11, 11, 11, 11, 40, 70]

    summary_rows = []

    for sheet_label, desc_rel, res_key, child_keys in ENDPOINTS:
        desc_path = os.path.join(BASE, desc_rel)
        endpoint_name, attrs = extract_attrs(desc_path, res_key, child_keys)

        # Filter to queryable only? No — show all, mark queryable
        # Sort: queryable first, then alphabetical
        attrs_sorted = sorted(attrs, key=lambda a: (not a.get('queryable', False), a.get('name', '')))

        queryable_count = sum(1 for a in attrs if a.get('queryable', False))
        total_count = len(attrs)

        ws = wb.create_sheet(title=sheet_label)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        if child_keys:
            title_text = f"{res_key} → {child_keys[0]}  ({queryable_count} queryable / {total_count} total)"
        else:
            title_text = f"{res_key}  ({queryable_count} queryable / {total_count} total)"
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = Font(name='Calibri', bold=True, size=14, color='2F5496')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 28

        # Header row
        for ci, hdr in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=hdr)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        ws.row_dimensions[2].height = 22

        # Data rows
        for ri, attr in enumerate(attrs_sorted, 3):
            name = attr.get('name', '')
            atype = attr.get('type', '')
            queryable = attr.get('queryable', False)
            updatable = attr.get('updatable', False)
            mandatory = attr.get('mandatory', False)
            max_len = attr.get('maxLength', attr.get('precision', ''))
            title = attr.get('title', '')
            desc = ''
            annotations = attr.get('annotations', {})
            if annotations:
                desc = annotations.get('description', '')

            row_data = [
                ri - 2,
                name,
                atype,
                'Yes' if queryable else 'No',
                'Yes' if updatable else 'No',
                'Yes' if mandatory else 'No',
                str(max_len) if max_len else '',
                title,
                desc,
            ]

            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGN
                cell.border = THIN_BORDER
                if (ri - 3) % 2 == 1:
                    cell.fill = ALT_FILL

                # Color-code Yes/No columns
                if ci in (4, 5, 6):
                    cell.font = YES_FONT if val == 'Yes' else NO_FONT

        # Column widths
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=2, column=ci).column_letter].width = w

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(headers)).column_letter}{2 + len(attrs_sorted)}"

        summary_rows.append((sheet_label, endpoint_name, child_keys, queryable_count, total_count))

    # --- Summary sheet ---
    summary = wb.create_sheet(title="Summary", index=0)
    summary.merge_cells('A1:F1')
    t = summary.cell(row=1, column=1, value="Fusion REST API — Queryable Fields Catalog")
    t.font = Font(name='Calibri', bold=True, size=16, color='2F5496')
    t.alignment = Alignment(horizontal='left')
    summary.row_dimensions[1].height = 30

    sum_headers = ['Sheet', 'Resource', 'Child', 'Queryable Fields', 'Total Fields', '% Queryable']
    for ci, h in enumerate(sum_headers, 1):
        c = summary.cell(row=3, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER

    total_q = 0
    total_t = 0
    for ri, (label, ename, ckeys, qcount, tcount) in enumerate(summary_rows, 4):
        total_q += qcount
        total_t += tcount
        pct = f"{qcount/tcount*100:.0f}%" if tcount else ""
        row_data = [label, ename, ', '.join(ckeys) if ckeys else '(parent)', qcount, tcount, pct]
        for ci, val in enumerate(row_data, 1):
            c = summary.cell(row=ri, column=ci, value=val)
            c.font = CELL_FONT
            c.alignment = CELL_ALIGN
            c.border = THIN_BORDER
            if (ri - 4) % 2 == 1:
                c.fill = ALT_FILL

    # Totals
    tr = 4 + len(summary_rows)
    for ci, val in enumerate(['TOTAL', '', '', total_q, total_t,
                               f"{total_q/total_t*100:.0f}%" if total_t else ""], 1):
        c = summary.cell(row=tr, column=ci, value=val)
        c.font = Font(name='Calibri', bold=True, size=11)
        c.border = THIN_BORDER

    summary.column_dimensions['A'].width = 22
    summary.column_dimensions['B'].width = 28
    summary.column_dimensions['C'].width = 20
    summary.column_dimensions['D'].width = 16
    summary.column_dimensions['E'].width = 14
    summary.column_dimensions['F'].width = 14
    summary.freeze_panes = 'A4'

    return wb


if __name__ == '__main__':
    wb = build_workbook()
    out = os.path.join(BASE, 'Queryable_Fields_Catalog.xlsx')
    wb.save(out)
    print(f"Saved: {out}")
    for ws in wb.worksheets:
        if ws.title == 'Summary':
            print(f"  Sheet '{ws.title}'")
        else:
            print(f"  Sheet '{ws.title}': {ws.max_row - 2} fields")
